from django.shortcuts import render

# Create your views here.
import pandas as pd
from io import BytesIO
import openpyxl
from django.http import HttpResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from xlsx_export.number_to_text import num2text


@csrf_exempt
def attendance_form(request):
    num_days = 7  # or any other default number

    if request.method == "POST":
        all_people = set()
        attendance_lists = []

        for i in range(1, num_days + 1):
            names = request.POST.get(f"day{i}", "")
            people = set(n.strip() for n in names.strip().splitlines() if n.strip())
            attendance_lists.append((f"Day {i}", people))
            all_people.update(people)

        all_people = sorted(all_people)
        data = {"Name": all_people}
        # First create numeric data for calculations
        numeric_data = {"Name": all_people}
        for label, people in attendance_lists:
            numeric_data[label] = [3 if name in people else 0 for name in all_people]
            data[label] = [3 if name in people else '-' for name in all_people]

        # Create DataFrame with symbols and calculate total using numeric data
        df = pd.DataFrame(data)
        numeric_df = pd.DataFrame(numeric_data)

        # Calculate totals and convert to text
        row_totals = numeric_df.iloc[:, 1:].sum(axis=1)
        df['Total'] = [num2text(int(total), )
                       for total in row_totals]

        # Calculate daily totals and grand total
        daily_totals = [numeric_df[col].sum() for col in numeric_df.columns[1:]]
        grand_total = sum(daily_totals)

        # Create complete totals row including the grand total
        totals = ['Всього'] + daily_totals + [num2text(int(grand_total),
                                                       main_units=((u'година', u'години', u'годин'), 'f'))]
        df.loc[len(df)] = totals  # Add the totals row to the display DataFrame

        output = BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)

        # Get the workbook from the BytesIO object
        wb = openpyxl.load_workbook(output)
        ws = wb.active

        # Set column widths
        ws.column_dimensions['A'].width = 35  # Name column
        # Total column width for text
        ws.column_dimensions[get_column_letter(ws.max_column)].width = 50

        # Define styles
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
        center_alignment = Alignment(horizontal='center')

        # Apply formatting to cells with numbers
        last_row = len(list(ws.rows))
        for row in ws.iter_rows(min_row=2):  # Start from row 2 to skip header
            row_num = row[0].row
            is_total_row = (row_num == last_row)

            # Format the first cell of totals row
            if is_total_row:
                row[0].font = openpyxl.styles.Font(bold=True)
                row[0].alignment = Alignment(horizontal='right')

            # Format day columns
            for cell in row[1:-1]:  # Skip the name column and total column
                cell.alignment = center_alignment
                if is_total_row:
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9',
                                            fill_type='solid')  # Very light green
                else:
                    if cell.value == 3:
                        cell.fill = green_fill
                    elif cell.value == '-':
                        cell.fill = red_fill

            # Format total cell
            total_cell = row[-1]  # Last cell in the row
            total_cell.alignment = Alignment(
                horizontal='left', wrap_text=True)  # Left align and wrap text
            if is_total_row:
                total_cell.font = openpyxl.styles.Font(bold=True)
                total_cell.fill = PatternFill(
                    start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')  # Medium light green
            else:
                total_cell.fill = PatternFill(
                    start_color='B3E6B3', end_color='B3E6B3', fill_type='solid')  # Light green

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=attendance.xlsx"
        return response

    return render(request, "prod_handout/attendance_form.html", {"range": range(1, num_days + 1)})
