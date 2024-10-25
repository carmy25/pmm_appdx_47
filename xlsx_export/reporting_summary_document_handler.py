from openpyxl.utils import get_column_letter

from receipts.models.reporting import FALReportEntry
from .utils import CENTER_ALIGNMENT, THIN_BORDER, cell_center_border
from .base_document_handler import BaseFALDocumentHandler


class ReportingSummaryReportDocumentHandler(BaseFALDocumentHandler):
    """Handle ReportingSummaryReport"""

    def update_total_base_dep(self):
        pass

    def get_document_name(self):
        return self.fal.report.summary_report._meta.verbose_name

    def get_dep(self):
        return ""

    def get_document_number(self):
        return self.fal.report.summary_report.number or "-"

    def get_document_operation_date(self):
        return self.fal.report.summary_report.end_date

    def get_document_sender(self):
        return "списання"

    def get_fal_income(self):
        return 0

    def get_fal_outcome(self):
        fals = FALReportEntry.objects.filter(
            fal_type=self.fal.fal_type,
            report__summary_report=self.fal.report.summary_report,
        )
        return sum([fal.outcome for fal in fals])

    def process(self):
        if (sr := self.fal.report.summary_report) and (
            sr in self.state.setdefault("reports_processed", [])
        ):
            return False
        self.state["reports_processed"].append(self.fal.report.summary_report)
        super().process()
        self.format_departments()
        return True

    def format_departments(self):
        fals = FALReportEntry.objects.filter(
            fal_type=self.fal.fal_type,
            report__summary_report=self.fal.report.summary_report,
        )
        for fal in fals:
            dep_index = self.state["DEP_BY_INDEX"][fal.report.department.name]
            cell_center_border(
                self.ws, self.add_idx(get_column_letter(dep_index + 1)), fal.outcome
            )
            col_letter = get_column_letter(dep_index + 2)
            total_cell = self.add_idx(col_letter)
            self.ws[total_cell].value = (
                f'=SUM({get_column_letter(dep_index)}$3:{get_column_letter(dep_index)}{self.state["idx"]})-SUM({
                get_column_letter(dep_index+1)}$3:{get_column_letter(dep_index+1)}{self.state["idx"]})'
            )
            self.ws[total_cell].alignment = CENTER_ALIGNMENT
            self.ws[total_cell].border = THIN_BORDER
