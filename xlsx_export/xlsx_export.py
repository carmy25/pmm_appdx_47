from collections import defaultdict
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from itertools import chain
from fals.models import FAL
from receipts.models.handout_list import HandoutList
from receipts.models.inspection_certificate import InspectionCertificate
from receipts.models.invoice import Invoice
from receipts.models.reporting import FALReportEntry
from receipts.models import ReceiptRequest, ReceiptRequestCoupon, Certificate
from receipts.models.writing_off_act import WritingOffAct
from xlsx_export.handout_list_summary_document_handler import HandoutSummaryDocumentHandler
from xlsx_export.invoice_summary_document_handler import InvoiceSummaryReportDocumentHandler
from xlsx_export.invoices_for_rrc import report_price_format_fals, report_price_format_header

from .utils import cell_center_border, THIN_BORDER
from .reporting_summary_document_handler import ReportingSummaryReportDocumentHandler
from .base_document_handler import BaseFALDocumentHandler
from .invoice_document_handler import InvoiceDocumentHandler

import logging
logger = logging.getLogger(__name__)
DEP_BY_INDEX = {}


class FALDocumentHandler(BaseFALDocumentHandler):

    def get_dep(self):
        if type(self.fal.document_object) in [ReceiptRequestCoupon, Certificate]:
            return self.fal.document_object.destination
        return self.fal.document_object.sender

    def update_total_base_dep(self):
        if type(self.fal.document_object) in [ReceiptRequestCoupon, Certificate]:
            destination = self.fal.document_object.destination
            self.state["total_by_dep"][destination] = self.fal.amount + self.state[
                "total_by_dep"
            ].get(destination, 0)
        else:
            sender = self.fal.document_object.sender
            self.state["total_by_dep"][sender] = (
                self.state["total_by_dep"].get(sender, 0) - self.fal.amount
            )

    def get_document_name(self):
        return self.fal.document_object._meta.verbose_name

    def get_document_number(self):
        doc = self.fal.document_object
        if type(doc) in [ReceiptRequest, ReceiptRequestCoupon]:
            return f"{
                doc.number}/{doc.book_number}{doc.book_series.upper()}"
        return doc.number

    def get_fal_income(self):
        doc = self.fal.document_object
        if type(doc) in [Certificate, ReceiptRequestCoupon]:
            return self.fal.amount
        return 0

    def get_fal_outcome(self):
        doc = self.fal.document_object
        if type(doc) in [Certificate, ReceiptRequestCoupon]:
            return 0
        return self.fal.amount

    def get_document_sender(self):
        doc = self.fal.document_object
        if type(doc) in [Certificate, ReceiptRequestCoupon]:
            return doc.sender.upper()
        return doc.destination.upper()

    def get_document_operation_date(self):
        return self.fal.document_object.operation_date


def export_fal_type(fal_type, ws, departments):
    format_header(ws, fal_type, departments)
    format_rows(ws, fal_type)


def registry_fes_format_header(ws, date):
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["f"].width = 30
    ws.row_dimensions[1].height = 30
    ws.merged_cells.ranges.add('A1:f1')
    year, month = date
    c = cell_center_border(ws, 'A1', f'Реєстр документів для ФЕС за {month}-{year}')
    c.font = Font(bold=True, size=20)
    c.alignment = Alignment(vertical='center', horizontal='center')

    c = cell_center_border(ws, 'A2', f'номери зведених')
    c.font = Font(bold=True)
    ws.merged_cells.ranges.add('b2:f2')
    c = cell_center_border(ws, 'B2', f'')
    c = cell_center_border(ws, 'c2', f'')
    c = cell_center_border(ws, 'd2', f'')
    c = cell_center_border(ws, 'e2', f'')
    c = cell_center_border(ws, 'f2', f'')

    c = cell_center_border(ws, 'A3', f'номери накладних')
    c.font = Font(bold=True)
    ws.merged_cells.ranges.add('b3:f3')
    c = cell_center_border(ws, 'B3', f'')
    c = cell_center_border(ws, 'c3', f'')
    c = cell_center_border(ws, 'd3', f'')
    c = cell_center_border(ws, 'e3', f'')
    c = cell_center_border(ws, 'f3', f'')

    c = cell_center_border(ws, 'A4', f'№')
    c.font = Font(bold=True)

    c = cell_center_border(ws, 'b4', f'номер донесення')
    c.font = Font(bold=True)
    c = cell_center_border(ws, 'c4', f'назва підрозділу')
    c.font = Font(bold=True)
    c = cell_center_border(ws, 'd4', f'дата')
    c.font = Font(bold=True)
    c = cell_center_border(ws, 'e4', f'номери шляхових')
    c.font = Font(bold=True)
    c = cell_center_border(ws, 'f4', f'номери роздавалних відомостей')
    c.font = Font(bold=True)


def comma_join(s):
    if s is None:
        return ''
    return ', '.join(s.split())


def registry_fes_format_rows(ws, reportings):
    for i, reporting in enumerate(reportings, 5):
        cell_center_border(ws, f'a{i}', i - 4)
        cell_center_border(ws, f'b{i}', reporting.number)
        cell_center_border(ws, f'C{i}', reporting.department.name)
        c = cell_center_border(ws, f'd{i}', reporting.end_date)
        c.number_format = 'DD/MM/YY'
        cell_center_border(ws, f'e{i}', comma_join(reporting.waybills_numbers))
        cell_center_border(ws, f'f{i}', comma_join(reporting.handout_numbers))

    return i


def registry_fes_format_footer(ws, last_idx):
    idx = last_idx + 1

    idx += 4
    ws[f'A{idx}'] = 'Здав:'
    ws[f'A{idx}'].font = Font(bold=True, size=16)
    ws[f'B{idx}'] = '__________________________' * 3
    ws.merged_cells.ranges.add(f'B{idx}:D{idx}')

    idx += 4
    ws[f'A{idx}'] = 'Прийняв:'
    ws[f'A{idx}'].font = Font(bold=True, size=14)
    ws[f'B{idx}'] = '__________________________' * 3
    ws.merged_cells.ranges.add(f'B{idx}:D{idx}')


def export_reportings_fes_registry(ws, reportings, date):
    registry_fes_format_header(ws, date)
    last_idx = registry_fes_format_rows(ws, reportings)
    registry_fes_format_footer(ws, last_idx)


def get_fal_date(obj):
    if isinstance(obj, FALReportEntry):
        return obj.report.summary_report.end_date
    return obj.document_object.operation_date


def get_sorted_fals(fal_type):
    fals = (
        FAL.objects.filter(fal_type=fal_type)
        .exclude(object_id__isnull=True)
        .prefetch_related("document_object")
    )
    fal_report_entries = (
        FALReportEntry.objects.filter(fal_type=fal_type)
        .exclude(report__isnull=True)
        .exclude(report__summary_report__isnull=True)
    )

    fals = filter(lambda x: not isinstance(
        getattr(x, 'document_object', None) or x.report, (WritingOffAct, InspectionCertificate)), chain(fals, fal_report_entries))
    sorted_fals = sorted(fals, key=get_fal_date)
    return sorted_fals


def format_rows(ws, fal_type):
    ws_state = {"total": 0, 'reports_processed': [],
                "total_by_dep": {}, "DEP_BY_INDEX": DEP_BY_INDEX}
    fals = get_sorted_fals(fal_type)
    j = 3
    for i, fal in enumerate(fals):
        ws_state["idx"] = i + j
        if type(fal) is FALReportEntry:
            if not ReportingSummaryReportDocumentHandler(fal, ws, ws_state).process():
                j -= 1
        elif isinstance(fal.document_object, WritingOffAct):
            if not HandoutSummaryDocumentHandler(fal, ws, ws_state).process():
                j -= 1
        elif isinstance(fal.document_object, HandoutList):
            if fal.document_object.summary_report is None:
                j -= 1
                continue
            if not HandoutSummaryDocumentHandler(fal, ws, ws_state).process():
                j -= 1
        elif type(fal.document_object) is not Invoice:
            FALDocumentHandler(fal, ws, ws_state).process()
        elif type(fal.document_object) is Invoice:
            if fal.document_object.summary_report:
                if not InvoiceSummaryReportDocumentHandler(fal, ws, ws_state).process():
                    j -= 1
            else:
                InvoiceDocumentHandler(fal, ws, ws_state).process()


def format_header(ws, fal_type, departments):
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["b"].width = 30
    ws.column_dimensions["c"].width = 30
    ws.column_dimensions["d"].width = 30
    ws.column_dimensions["e"].width = 20
    ws.column_dimensions["f"].width = 20
    ws.column_dimensions["g"].width = 20
    ws["A1"] = fal_type.name
    ws["a1"].border = THIN_BORDER
    ws["a1"].alignment = Alignment(horizontal="center")
    ws["a1"].font = Font(bold=True)
    ws["a1"].fill = PatternFill(
        start_color="c4d79b", end_color="c4d79b", fill_type="solid"
    )
    ws.merged_cells.ranges.add("A1:D1")
    ws.merged_cells.ranges.add("E1:G1")
    total_cell = cell_center_border(ws, "E1", "Загалом")
    total_cell.font = Font(bold=True)
    total_cell.fill = PatternFill(
        start_color="9bb8d9", end_color="9bb8d9", fill_type="solid"
    )

    cell_center_border(
        ws, "a2", "Найменування документу").font = Font(bold=True)
    cell_center_border(ws, "b2", "Номер документу").font = Font(bold=True)
    cell_center_border(ws, "c2", "Дата операції").font = Font(bold=True)
    cell_center_border(
        ws, "d2", "Постачальник (одержувач)").font = Font(bold=True)
    cell_center_border(ws, "e2", "Надійшло").font = Font(bold=True)
    cell_center_border(ws, "f2", "Вибуло").font = Font(bold=True)
    cell_center_border(ws, "g2", "Всього").font = Font(bold=True)

    format_departments(ws, departments)


def format_departments(ws, deps):
    col_idx = 8
    for dep in deps:
        DEP_BY_INDEX[dep.name] = col_idx
        start_cell_name = f"{get_column_letter(col_idx)}1"
        ws.merged_cells.ranges.add(
            f"{start_cell_name}:{
                get_column_letter(col_idx+2)}1"
        )
        dep_name_cell = cell_center_border(ws, start_cell_name, dep.name)
        dep_name_cell.font = Font(bold=True)
        dep_name_cell.fill = PatternFill(
            start_color="fee8d9", end_color="fee8d9", fill_type="solid"
        )

        cell_center_border(
            ws,
            f"{get_column_letter(
                col_idx)}2",
            "Надійшло",
        ).font = Font(bold=True)
        cell_center_border(
            ws,
            f"{get_column_letter(
                col_idx+1)}2",
            "Вибуло",
        ).font = Font(bold=True)
        cell_center_border(
            ws,
            f"{get_column_letter(
                col_idx+2)}2",
            "Всього",
        ).font = Font(bold=True)
        col_idx += 3


def export_reportings_price_report(ws, reportings, date, invoices, start_date):
    # fal_types_by_idx = {ft: i for i, ft in enumerate({r.fal_type for r in reportings})}
    fal_types_amount = defaultdict(float)
    fal_types_total_amount = defaultdict(float)
    fal_types_price = defaultdict(float)
    for report in reportings:
        next_fre = False
        for fre in report.fals.all():
            # decrement outcome from invoices
            # accumulated sum
            outcome = fre.get_outcome_kgs()
            write_off_total = 0
            write_off_price = 0
            invoices_range = invoices.loc[start_date:report.end_date]
            next_fre = False
            for invoice in reversed(invoices_range['invoices'].tolist()):
                if next_fre:
                    break
                for fal in invoice.fals:
                    if fal.fal_type.get_name() == fre.fal_type.get_name():
                        write_off = fal.write_off(outcome-write_off_total)
                        write_off_total += write_off['amount']
                        write_off_price += write_off['price']
                        if write_off['price'] and write_off['amount']:
                            if (write_off['price']/write_off['amount'] > 60):
                                logger.warning(f'Invoice({fal.fal_type}) {
                                    fal.doc.invoice_for_rrc.number} looks incorrect')
                        if write_off_total >= outcome:
                            next_fre = True
                            break
            fal_types_amount[fre.fal_type.name] += write_off_total
            fal_types_price[fre.fal_type.name] += write_off_price
            fal_types_total_amount[fre.fal_type.name] += outcome
    report_price_format_header(ws, date)
    report_price_format_fals(ws, fal_types_amount, fal_types_price, fal_types_total_amount)
