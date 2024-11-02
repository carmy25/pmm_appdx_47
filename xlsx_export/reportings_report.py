from datetime import timedelta
from openpyxl.utils import get_column_letter
from collections import Counter
from openpyxl.comments import Comment
from django.core.exceptions import ObjectDoesNotExist
from django.core.exceptions import MultipleObjectsReturned

from receipts.exceptions import DuplicateReportsError
from xlsx_export.utils import cell_center_border, month_iter


MONTH_BY_INDEX = {}
DEP_BY_INDEX = {}


def format_departments_column(ws, deps):
    idx = 2
    for dep in set(deps):
        DEP_BY_INDEX[dep.name] = idx
        cell_center_border(ws, f"A{idx}", dep.name)
        idx += 1


def export_reportings_report(ws, reportings):
    departments = [r.department for r in reportings]
    ws.column_dimensions["A"].width = 40
    format_departments_column(ws, departments)
    oldest = reportings.first().end_date
    newest = reportings.last().end_date

    idx = 2
    for y, m in month_iter(oldest.month, oldest.year, newest.month, newest.year):
        col_name = get_column_letter(idx)
        MONTH_BY_INDEX[(y, m)] = col_name
        idx += 1
        cell_center_border(ws, f"{col_name}1", f"{m}/{y}")

    format_deps_reportings(ws, reportings)
    ws.freeze_panes = ws["B2"]


def get_prev_report(report, all_reports):
    start_date = report.start_date
    for r in all_reports:
        if (
            r.end_date == start_date - timedelta(days=1)
            and r.department == report.department
        ):
            return r
    return None


def format_deps_reportings(ws, reportings):
    for report in reportings:
        state_field = report._meta.get_field('state')
        display_value = report._get_FIELD_display(state_field)
        comment = f"Стан: {display_value}]n"
        good = True
        end_date = report.end_date
        col_name = MONTH_BY_INDEX[(end_date.year, end_date.month)]
        row_idx = DEP_BY_INDEX[report.department.name]

        prev_reporting = get_prev_report(report, reportings)
        if prev_reporting is None:
            comment = "Попереднє донесення відсутнє\n"
        cell_addr = f"{col_name}{row_idx}"
        for fal in report.fals.all():
            if fal.remains != 0:
                try:
                    old_fal = prev_reporting.fals.get(fal_type=fal.fal_type)
                    old_fal_all = round(
                        old_fal.remains + old_fal.income - old_fal.outcome, 1
                    )
                except AttributeError:
                    pass
                except ObjectDoesNotExist:
                    good = False
                    comment += f"{fal.fal_type.name} відсутнє\n"
                except MultipleObjectsReturned as exc:
                    duplicates = [
                        item
                        for item, count in Counter(
                            [f.fal_type.name for f in prev_reporting.fals.all()]
                        ).items()
                        if count > 1
                    ]
                    raise DuplicateReportsError(
                        f"В донесенні [{prev_reporting.department.name}] за [{
                            prev_reporting.start_date} - {prev_reporting.end_date}] дублюється [{duplicates}]"
                    ) from exc
                else:
                    if fal.remains != old_fal_all:
                        good = False
                        comment += (
                            f"{fal.fal_type.name} {fal.remains} <> {old_fal_all}\n"
                        )
            density = fal.get_density()
            income_kgs = round(fal.income * density, 1)
            remains_kgs = round(fal.remains * density, 1)
            outcome_kgs = round(fal.outcome * density, 1)
            total_kgs = remains_kgs + income_kgs - outcome_kgs
            comment += f"{fal.fal_type.name}({remains_kgs:.1f}/{income_kgs:.1f}/{
                outcome_kgs:.1f}) - {total_kgs:.1f}\n\n"
        if report.fals.all().count() == 0:
            good = False
        cell_center_border(ws, cell_addr, "+" if good else "-")
        ws[cell_addr].comment = Comment(
            comment + (report.note if report.note else ""),
            "auth",
            height=200,
            width=350,
        )
