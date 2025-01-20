from openpyxl.utils import get_column_letter
from collections import defaultdict

from .utils import CENTER_ALIGNMENT, THIN_BORDER, cell_center_border
from .base_document_handler import BaseFALDocumentHandler


class HandoutSummaryDocumentHandler(BaseFALDocumentHandler):
    ''''Process handoff list summary reports.'''

    def update_total_base_dep(self):
        pass

    def get_dep(self):
        return ""

    def get_document_name(self):
        sr = self.fal.document_object.summary_report
        return sr._meta.verbose_name

    def get_dep(self):
        return ""

    def get_document_number(self):
        sr = self.fal.document_object.summary_report
        return sr.number

    def get_document_operation_date(self):
        sr = self.fal.document_object.summary_report
        return sr.document_date

    def get_document_sender(self):
        return "в підр.(зведена розд.)"

    def get_fal_income(self):
        return 0

    def get_fal_outcome(self):
        return 0

    def process(self):
        if (sr := self.fal.document_object.summary_report) and (
            sr in self.state.setdefault("reports_processed", [])
        ):
            return False
        self.state["reports_processed"].append(self.fal.document_object.summary_report)
        super().process()
        self.format_departments()
        return True

    def format_dep_fal(self, dep, amount, income=True):
        dep_index = self.state["DEP_BY_INDEX"][dep.name]
        cell_center_border(
            self.ws,
            self.add_idx(get_column_letter(dep_index if income else dep_index + 1)),
            amount,
        )
        col_letter = get_column_letter(dep_index + 2)
        total_cell = self.add_idx(col_letter)
        self.ws[total_cell].value = (
            f'=SUM({get_column_letter(dep_index)}$3:{
                get_column_letter(dep_index)}{self.state["idx"]})-SUM({
                get_column_letter(dep_index+1)}$3:{get_column_letter(dep_index+1)}{
                self.state["idx"]})'
        )
        self.ws[total_cell].alignment = CENTER_ALIGNMENT
        self.ws[total_cell].border = THIN_BORDER

    def format_departments(self):
        sr = self.fal.document_object.summary_report
        docs = sr.documents.filter(fals__fal_type=self.fal.fal_type)
        sender_amounts = defaultdict(int)
        dst_amounts = defaultdict(int)
        for doc in docs:
            amount = doc.fals.filter(fal_type=self.fal.fal_type).first().amount

            sender_amounts[doc.sender] += amount
            dst_amounts[doc.destination] += amount

        for d, a in sender_amounts.items():
            self.format_dep_fal(d, a, income=False)
        for d, a in dst_amounts.items():
            self.format_dep_fal(d, a)
