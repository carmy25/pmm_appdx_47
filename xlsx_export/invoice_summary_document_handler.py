from openpyxl.utils import get_column_letter

from .utils import CENTER_ALIGNMENT, THIN_BORDER, cell_center_border
from .base_document_handler import BaseFALDocumentHandler


class InvoiceSummaryReportDocumentHandler(BaseFALDocumentHandler):
    """Process invoices summary report"""

    def update_total_base_dep(self):
        pass

    def get_dep(self):
        return ""

    def get_document_name(self):
        sr = self.fal.document_object.summary_report
        return sr._meta.verbose_name

    def get_dep(self):
        return ""

    def get_document_number(self):
        sr = self.fal.document_object.summary_report
        return sr.number

    def get_document_operation_date(self):
        sr = self.fal.document_object.summary_report
        return sr.document_date

    def get_document_sender(self):
        return "в підр.(зведена)"

    def get_fal_income(self):
        return 0

    def get_fal_outcome(self):
        return 0

    def process(self):
        if (sr := self.fal.document_object.summary_report) and (
            sr in self.state.setdefault("reports_processed", [])
        ):
            return False
        self.state.setdefault("reports_processed", []).append(
            self.fal.document_object.summary_report
        )
        super().process()
        self.format_departments()
        return True

    def format_dep_fal(self, dep, amount, income=True):
        dep_index = self.state["DEP_BY_INDEX"][dep.name]
        cell_center_border(
            self.ws,
            self.add_idx(get_column_letter(dep_index if income else dep_index + 1)),
            amount,
        )
        col_letter = get_column_letter(dep_index + 2)
        total_cell = self.add_idx(col_letter)
        self.ws[total_cell].value = (
            f'=SUM({get_column_letter(dep_index)}$3:{
                get_column_letter(dep_index)}{self.state["idx"]})-SUM({
                get_column_letter(dep_index+1)}$3:{get_column_letter(dep_index+1)}{
                self.state["idx"]})'
        )
        self.ws[total_cell].alignment = CENTER_ALIGNMENT
        self.ws[total_cell].border = THIN_BORDER

    def format_departments(self):
        sr = self.fal.document_object.summary_report
        docs = sr.invoices.filter(fals__fal_type=self.fal.fal_type)
        sender_amounts = {}
        dst_amounts = {}
        for doc in docs:
            amount = doc.fals.filter(fal_type=self.fal.fal_type).first().amount

            if doc.sender not in sender_amounts:
                sender_amounts[doc.sender] = amount
            else:
                sender_amounts[doc.sender] += amount

            if doc.destination not in dst_amounts:
                dst_amounts[doc.destination] = amount
            else:
                dst_amounts[doc.destination] += amount

        for d, a in sender_amounts.items():
            self.format_dep_fal(d, a, income=False)
        for d, a in dst_amounts.items():
            self.format_dep_fal(d, a)
