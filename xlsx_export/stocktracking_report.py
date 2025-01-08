from copy import copy
from datetime import date, datetime
from django.conf import settings
import openpyxl as xl

from fals.models import FAL, FALType
from receipts.models.certificate import Certificate
from receipts.models.handout_list import HandoutList
from receipts.models.invoice import Invoice
from receipts.models.receipt import ReceiptRequest, ReceiptRequestCoupon
from receipts.models.reporting import FALReportEntry, Reporting
from xlsx_export.utils import cell_center_border
from .number_to_text import num2text


def format_dep_fals(ws, dep, data):
    i = 0
    kgs_total = 0
    for i, (fal_type, value) in enumerate(data['fals'].items(), 1):
        idx = value['idx']
        amount = value.setdefault('invoices_kgs', 0) + value.setdefault('handout_kgs',
                                                                        0) + value.setdefault('reporting_remains', 0)
        cell_center_border(ws, f'A{value["idx"]}', '=ROW()-35')
        cell_center_border(ws, f'B{value["idx"]}', '')
        cell_center_border(ws, f'C{value["idx"]}', fal_type.name)
        cell_center_border(ws, f'D{value["idx"]}', '')
        cell_center_border(ws, f'e{value["idx"]}', 'кг')
        cell_center_border(ws, f'F{value["idx"]}', amount)
        kgs_total += amount
        cell_center_border(ws, f'G{value["idx"]}', value['price'])
        cell_center_border(ws, f'H{value["idx"]}', f'=G{idx}*F{idx}')

        cell_center_border(ws, f'I{value["idx"]}', amount)
        cell_center_border(ws, f'J{value["idx"]}', value['price'])
        cell_center_border(ws, f'K{value["idx"]}', f'=G{idx}*F{idx}')
        cell_center_border(ws, f'L{value["idx"]}', '')
        cell_center_border(ws, f'm{value["idx"]}', '')
    if i == 0:
        return kgs_total
    cell_center_border(ws, f'F{36+i}', f'=sum(F36:F{35+i})')
    cell_center_border(ws, f'H{36+i}', f'=sum(H36:H{35+i})')
    cell_center_border(ws, f'I{36+i}', f'=sum(I36:I{35+i})')
    cell_center_border(ws, f'K{36+i}', f'=sum(K36:K{35+i})')
    return kgs_total


def format_dep_footer(ws, dep, data, kgs_total, form_data):
    ws_tmpl = settings.WB_TMPL.active
    entries_num = len(data['fals'].keys())
    idx = entries_num + 36 + 3
    for i in range(24):
        for j in range(1, 14):
            c = ws.cell(row=i+idx, column=j)
            tc = ws_tmpl.cell(row=i+39, column=j)
            c.value = tc.value
            c.font = copy(tc.font)
            c.border = copy(tc.border)
            c.fill = copy(tc.fill)
            c.alignment = copy(tc.alignment)
    ws[f'c{idx}'].value = f'а) кількість порядкових номерів - {num2text(entries_num)}'
    ws[f'c{
        idx+2}'].value = f'б) загальна кількість кілограм,  фактично - {num2text(round(kgs_total))}'
    ws[f'c{
        idx+4}'].value = f'в) загальна кількість кілограм,  за даними бухгалтерського обліку - {num2text(round(kgs_total))}'
    # ws.row_dimensions[idx].height = 30
    ws.merged_cells.ranges.add(f'a{idx+17}:m{idx+18}')
    chief_pos, chief_name = get_position_and_name(form_data['committee_chief'])
    ws[f'C{idx+7}'].value = chief_pos
    ws[f'K{idx+7}'].value = chief_name
    for i, member in enumerate(form_data['committee_members'].split('\n')):
        pos, name = get_position_and_name(member)
        ws[f'C{idx+9 + i * 2}'].value = pos
        ws[f'K{idx+9 + i * 2}'].value = name


def get_position_and_name(rec):
    return map(lambda x: x.strip(), rec.split(';;'))


def export_stocktaking_report(wb, deps, form_data):
    data = {}
    for dep in deps:
        data[dep.name] = {}
        update_reporting_dep_data(dep, data[dep.name])
        update_receipts_dep_data(dep, data[dep.name])
        update_invoice_dep_data(dep, data[dep.name])
        update_handout_dep_data(dep, data[dep.name])
        update_price_dep_data(dep, data[dep.name])

        ws = wb.create_sheet(dep.name)
        format_dep_header(ws, dep, data[dep.name], form_data)
        kgs_total = format_dep_fals(ws, dep, data[dep.name])
        format_dep_footer(ws, dep, data[dep.name], kgs_total, form_data)

    ws = wb.create_sheet('total')
    format_total_sheet(ws, data)


def format_total_sheet(ws, data):
    idx = 1
    cell_center_border(ws, f'A{idx}', 'Вид ПММ')
    cell_center_border(ws, f'B{idx}', 'К-сть(кг)')
    cell_center_border(ws, f'C{idx}', 'Після Атестату')
    cell_center_border(ws, f'D{idx}', 'Після ТЧВ')
    cell_center_border(ws, f'E{idx}', 'Після ЧВ')
    cell_center_border(ws, f'F{idx}', 'Після Донесень')
    totals = {}
    idx += 1
    for fal_type in FALType.objects.all():
        name = fal_type.get_name()
        for certificate in Certificate.objects.all():
            try:
                fal = certificate.fals.get(fal_type__name=name)
                if totals.get(name):
                    totals[name] += fal.amount
                else:
                    totals[name] = fal.amount
            except FAL.DoesNotExist:
                pass
        try:
            cell_center_border(ws, f"C{idx}", totals[name])
        except:
            print(f'Not found: {name}')

        for rrc in ReceiptRequestCoupon.objects.all():
            try:
                fal = rrc.fals.get(fal_type__name=name)
                if totals.get(name):
                    totals[name] += fal.amount
                else:
                    totals[name] = fal.amount
            except FAL.DoesNotExist:
                pass

        try:
            cell_center_border(ws, f"D{idx}", totals[name])
        except:
            print(f'Not found: {name}')

        for rr in ReceiptRequest.objects.all():
            try:
                fal = rr.fals.get(fal_type__name=name)
                totals[name] -= fal.amount
            except FAL.DoesNotExist:
                pass

        try:
            cell_center_border(ws, f"E{idx}", totals[name])
        except:
            print(f'Not found: {name}')

        for reporting in Reporting.objects.all():
            try:
                fal = reporting.fals.get(fal_type__name=name)
                totals[name] -= fal.get_outcome_kgs()
            except FALReportEntry.DoesNotExist:
                pass
            except Exception as e:
                print(e)
        cell_center_border(ws, f"A{idx}", name)
        try:
            cell_center_border(ws, f"B{idx}", totals[name])
        except:
            print(f'Not found: {name}')
        idx += 1

    '''
    for dep_name, values in data.items():
        for fal_type, amounts in values.setdefault('fals', {}).items():
            if not totals.get(fal_type):
                totals[fal_type.name] = {'idx': 0, 'amount': 0}
    for dep_name, values in data.items():
        for fal_type, amounts in values['fals'].items():
            if totals[fal_type.name]['idx'] == 0:
                totals[fal_type.name]['amount'] += amounts['reporting_remains'] + \
                    amounts['invoices_kgs'] + amounts['handout_kgs']
                totals[fal_type.name]['idx'] = idx
                idx += 1
            cell_center_border(ws, f"A{totals[fal_type.name]['idx']}", fal_type.name)
            cell_center_border(ws, f"B{totals[fal_type.name]['idx']}",
                               totals[fal_type.name]['amount'])
    '''


def update_price_dep_data(dep, data):
    for fal_type, d in data['fals'].items():
        rrc = fal_type.fal_rrc_entries.filter(
            invoice_for_rrc__rrc__operation_date__lte=data.setdefault(
                'end_date',  datetime.now())).last()
        if rrc is None:
            print(f'FAL not found: {fal_type.name}')
            rrc = fal_type.fal_rrc_entries.filter().last()
            if rrc is None:
                price = 15
                amount = 1
            else:
                price = rrc.price
                amount = rrc.amount
        else:
            price = rrc.price
            amount = rrc.amount

        data['fals'][fal_type]['price'] = price / amount


def update_invoice_dep_data(dep, data):
    report_end_date = data.get('end_date')
    invoices = dep.invoices_receiver.all()
    invoices_outcome = dep.invoices_sender.all()
    if report_end_date:
        invoices = invoices.filter(
            operation_date__gt=report_end_date).order_by('operation_date')
        invoices_outcome = invoices_outcome.filter(
            operation_date__gt=report_end_date).order_by('operation_date')
    indexes = map(lambda k: data['fals'][k].get('idx', 0), data['fals'].keys())
    idx = max(35, *(list(indexes) + [0])) + 1
    for invoice in invoices:
        for fal in invoice.fals.all():
            fal_data = data['fals'].setdefault(fal.fal_type, {})
            fal_data.setdefault('invoices_kgs', 0)
            if not fal_data.get('idx'):
                fal_data['idx'] = idx
                idx += 1
            fal_data['invoices_kgs'] += fal.amount

    for invoice in invoices_outcome:
        for fal in invoice.fals.all():
            fal_data = data['fals'].setdefault(fal.fal_type, {})
            fal_data.setdefault('invoices_kgs', 0)
            if not fal_data.get('idx'):
                fal_data['idx'] = idx
                idx += 1

            fal_data['invoices_kgs'] -= fal.amount


def update_handout_dep_data(dep, data):
    handouts = dep.received_handouts.all()
    sent_handouts = dep.sent_handouts.all()
    if data.get('end_date'):
        handouts = dep.received_handouts.filter(
            operation_date__gte=data.get('end_date'))
        sent_handouts = dep.sent_handouts.filter(
            operation_date__gte=data.get('end_date'))
    indexes = map(lambda k: data['fals'][k].get('idx', 0), data['fals'].keys())
    idx = max(35, *(list(indexes) + [0])) + 1
    for handout in handouts:
        for fal in handout.fals.all():
            fal_data = data['fals'].setdefault(fal.fal_type, {})
            fal_data.setdefault('handout_kgs', 0)
            if not fal_data.get('idx'):
                fal_data['idx'] = idx
                idx += 1
            fal_data['handout_kgs'] += fal.amount

    for handout in sent_handouts:
        for fal in handout.fals.all():
            fal_data = data['fals'].setdefault(fal.fal_type, {})
            fal_data.setdefault('handout_kgs', 0)
            if not fal_data.get('idx'):
                fal_data['idx'] = idx
                idx += 1
            fal_data['handout_kgs'] -= fal.amount


def format_dep_header(ws, dep, data, form_data):
    ws_tmpl = settings.WB_TMPL.active
    ws.merged_cells.ranges = ws_tmpl.merged_cells.ranges
    for i in range(1, 36):
        for j in range(1, 14):
            c = ws.cell(row=i, column=j)
            tc = ws_tmpl.cell(row=i, column=j)
            c.value = tc.value
            c.font = copy(tc.font)
            c.fill = copy(tc.fill)
            c.border = copy(tc.border)
            c.alignment = copy(tc.alignment)
    ws['A3'].value = form_data['department_name']
    ws['A4'].value = f'Ідентифікаційний кол за ЄДРПО: {form_data['edrpo_code']}'
    ws['A10'].value = form_data['date']
    ws['A12'].value = form_data['document_date_number']
    ws['A18'].value = f"станом на {form_data['date_remains']}"
    ws['C27'].value = f'розпочата {form_data["start_date"]}'
    chief_verbose = dep.chief_position_verbose()
    ws['D24'].value = f'{chief_verbose} {dep.name}'
    ws['K24'].value = f'{dep.chief_name}'
    ws['C28'].value = f'закінчена {form_data["end_date"]}'
    ws.column_dimensions['c'].width = 50
    ws['c16'].value = dep.name


def update_receipts_dep_data(dep, data):
    # if dep.name not in ['А4548', 'А4635']:
    #    return
    receipt_requests = ReceiptRequest.objects.filter(sender=dep.name)
    receipt_coupons = ReceiptRequestCoupon.objects.filter(destination=dep.name)
    certificates = Certificate.objects.filter(destination=dep.name)
    if (ed := data.get('end_date')):
        receipt_requests = receipt_requests.filter(operation_date__gt=ed)
        receipt_coupons = receipt_coupons.filter(operation_date__gt=ed)
        certificates = certificates.filter(operation_date__gt=ed)
    indexes = map(lambda k: data['fals'][k].get('idx', 0), data['fals'].keys())
    idx = max(35, *(list(indexes) + [0])) + 1
    for rc in receipt_coupons:
        print(f'Processing {str(rc)}')
        for fal in rc.fals.all():
            fal_data = data['fals'].setdefault(fal.fal_type, {})
            fal_data.setdefault('invoices_kgs', 0)
            if not fal_data.get('idx'):
                fal_data['idx'] = idx
                idx += 1
            fal_data['invoices_kgs'] += fal.amount
    for c in certificates:
        print(f'Processing {str(c)}')
        for fal in c.fals.all():
            fal_data = data['fals'].setdefault(fal.fal_type, {})
            fal_data.setdefault('invoices_kgs', 0)
            if not fal_data.get('idx'):
                fal_data['idx'] = idx
                idx += 1
            fal_data['invoices_kgs'] += fal.amount

    for rr in receipt_requests:
        print(f'Processing {str(rr)}')
        for fal in rr.fals.all():
            fal_data = data['fals'].setdefault(fal.fal_type, {})
            fal_data.setdefault('invoices_kgs', 0)
            if not fal_data.get('idx'):
                fal_data['idx'] = idx
                idx += 1
            fal_data['invoices_kgs'] -= fal.amount


def update_reporting_dep_data(dep, data):
    data['fals'] = {}
    reporting = dep.reportings.all().order_by('end_date').last()
    if reporting is None:
        return
    j = 0
    data['start_date'] = reporting.start_date
    data['end_date'] = reporting.end_date
    for i, fe in enumerate(reporting.fals.all().order_by('fal_type__name'), 36):
        idx = i - j
        remains_after_kgs = fe.get_remains_after_kgs()
        if remains_after_kgs == 0:
            j += 1
            continue
        data['fals'][fe.fal_type] = {
            'idx': idx,
            'reporting_remains': remains_after_kgs
        }
        '''
            rrc = (fe.fal_type
                   .fal_rrc_entries
                   .order_by('invoice_for_rrc__rrc__operation_date')
                   .last())
            if rrc is None or rrc.invoice_for_rrc.rrc is None:
                print(f'HHH {fe.fal_type.name}')
                continue
            print(f'{fe.fal_type.name} {rrc.invoice_for_rrc.rrc.operation_date}')
        '''
