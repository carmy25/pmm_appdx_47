from openpyxl.utils import get_column_letter

from .utils import cell_center_border


class BaseFALDocumentHandler:
    def __init__(self, fal, ws, state):
        self.ws = ws
        self.fal = fal
        self.state = state

    def add_idx(self, col_name):
        return f"{col_name}{self.state['idx']}"

    def format_document_name(self, name):
        cell_center_border(self.ws, self.add_idx("A"), name)

    def format_document_number(self, number):
        cell_center_border(self.ws, self.add_idx("B"), number)

    def format_document_operation_date(self, date):
        cell_center_border(self.ws, self.add_idx("C"), date)

    def format_document_sender(self, sender):
        cell_center_border(self.ws, self.add_idx("D"), sender)

    def format_fal_income(self, amount):
        cell_center_border(self.ws, self.add_idx("E"), amount or "")

    def format_fal_income_dep(self, amount):
        cell_center_border(self.ws, self.add_idx("E"), amount or "")

    def format_fal_outcome(self, amount):
        cell_center_border(self.ws, self.add_idx("F"), amount or "")

    def format_fal_total(self):
        sum_str = f'=sum(e$3:e{
            self.state["idx"]})-sum(f$3:f{self.state["idx"]})'
        cell_center_border(self.ws, self.add_idx("G"), sum_str)

    def format_fal_by_dep(self):
        income = self.get_fal_income()
        outcome = self.get_fal_outcome()
        dep = self.get_dep()
        dep_index_base = self.state["DEP_BY_INDEX"]["А4548"]
        dep_index_505 = self.state["DEP_BY_INDEX"]["А4635"]
        if dep == "А4548":
            cell_center_border(
                self.ws, self.add_idx(get_column_letter(dep_index_base)), income or ""
            )
            cell_center_border(
                self.ws,
                self.add_idx(get_column_letter(dep_index_base + 1)),
                outcome or "",
            )
            cell_center_border(
                self.ws, self.add_idx(get_column_letter(dep_index_505)), ""
            )
            cell_center_border(
                self.ws, self.add_idx(get_column_letter(dep_index_505 + 1)), ""
            )
        elif dep == "А4635":
            cell_center_border(
                self.ws, self.add_idx(get_column_letter(dep_index_base)), ""
            )
            cell_center_border(
                self.ws, self.add_idx(get_column_letter(dep_index_base + 1)), ""
            )
            cell_center_border(
                self.ws, self.add_idx(get_column_letter(dep_index_505)), income or ""
            )
            cell_center_border(
                self.ws,
                self.add_idx(get_column_letter(dep_index_505 + 1)),
                outcome or "",
            )

    def format_fal_total_base_dep(self):
        dep_index_base = self.state["DEP_BY_INDEX"]["А4548"]
        dep_index_505 = self.state["DEP_BY_INDEX"]["А4635"]
        cell_center_border(
            self.ws,
            self.add_idx(get_column_letter(dep_index_base + 2)),
            self.state["total_by_dep"].setdefault("А4548", 0),
        )
        cell_center_border(
            self.ws,
            self.add_idx(get_column_letter(dep_index_505 + 2)),
            self.state["total_by_dep"].setdefault("А4635", 0),
        )

    def process(self):
        self.format_document_name(self.get_document_name())
        self.format_document_number(self.get_document_number())
        self.format_document_operation_date(self.get_document_operation_date())
        self.format_document_sender(self.get_document_sender())
        self.format_fal_income(self.get_fal_income())
        self.format_fal_outcome(self.get_fal_outcome())

        self.update_total()
        self.format_fal_total()

        self.update_total_base_dep()
        self.format_fal_total_base_dep()
        self.format_fal_by_dep()

    def update_total(self):
        income = self.get_fal_income() or 0
        outcome = self.get_fal_outcome() or 0
        self.state["total"] = self.state["total"] - outcome + income
