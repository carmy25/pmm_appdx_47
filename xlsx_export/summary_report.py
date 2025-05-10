from fals.models import Category
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from constance import config
from xlsx_export.formulas import xl_sum
from xlsx_export.utils import LEFT_ALIGNMENT, cell_center, cell_center_border, cell_vertical_border, fill_empty_border, merge_rows


def push_fal_with_priority(fal, fal_set):
    priority = 9
    if fal.fal_type.category == Category.DIESEL:
        priority = 1
    elif fal.fal_type.category == Category.PETROL:
        priority = 2
    fal_set.add((priority, fal.fal_type.get_name()))


def collect_data(obj):
    docs_data = {}
    fal_types = set()
    for doc in obj.documents.all():
        docs_data[doc.number] = {
            'fals': {},
            'document_date': doc.document_date or doc.end_date,
            'department': doc.department.name,
        }
        for fal in doc.fals.all():
            fal_type = fal.fal_type.get_name()
            push_fal_with_priority(fal, fal_types)
            if (ft := docs_data[doc.number]['fals'].get(fal_type)) is not None:
                ft['ltrs'] += fal.outcome_burned_ltr
                ft['kgs'] += fal.get_burned_kgs()
            else:
                docs_data[doc.number]['fals'][fal_type] = {
                    'ltrs': fal.outcome_burned_ltr,
                    'kgs': fal.get_burned_kgs(),
                }
    return docs_data, sorted(fal_types)


def export_summary_report(ws, obj):
    docs_data, fal_types = collect_data(obj)
    format_table_header(ws, fal_types)
    idx = format_rows(ws, docs_data, fal_types)
    last_letter = get_column_letter(len(fal_types) * 2 + 5)
    rng = f'F11:{last_letter}{idx-1}'
    fill_empty_border(ws, rng, value=0)
    fill_empty_border(ws, f'B{idx}:E{idx}')
    format_header(ws, obj, len(fal_types))
    format_footer(ws, idx, obj, fal_types)


def format_header(ws, obj, header_len):
    end_letter = get_column_letter(header_len * 2 + 5)
    date_fmt = '%d.%m.%Y'
    merge_rows(ws, 'A',  end_letter, idx=2)
    c = cell_center(
        ws,
        f'A2',
        f'ЗВЕДЕНА ВІДОМІСТЬ № {obj.number} від {obj.document_date.strftime(date_fmt)} року')
    c.font = Font(bold=True, size=14)

    merge_rows(ws, 'A', end_letter, idx=3)
    c = cell_center(
        ws,
        f'A3',
        f'по донесенням про наявність та рух   пально-мастильних матеріалів у в/ч А4548')

    merge_rows(ws, 'A', end_letter, idx=4)
    c = cell_center(
        ws,
        f'A4',
        f'за період з {obj.start_date.strftime(date_fmt)} р. по {obj.end_date.strftime(date_fmt)} р.')


def format_footer(ws, idx, obj, fal_types):
    fal_column = 6
    fal_names = [name for _, name in fal_types]
    c = cell_center_border(ws, f'A{idx}', 'Усього за ПЗП :')
    c.font = Font(bold=True)
    c.alignment = LEFT_ALIGNMENT
    merge_rows(ws, 'A', 'E', idx)
    for i in range(fal_column, len(fal_names)*2):
        col_letter = get_column_letter(i)
        c = cell_center_border(
            ws,
            f'{col_letter}{idx}',
            str(xl_sum(f'{col_letter}11',
                       f'{col_letter}{idx-1}')))
        c.font = Font(bold=True)
    idx += 2
    footer_text = 'Начальник служби пального та мастильних матеріалів військової частини А4548'
    c = cell_center(ws, f'A{idx}', footer_text)
    c.alignment = LEFT_ALIGNMENT
    c.font = Font(size=14)
    merge_rows(ws, 'A', 'K', idx)

    idx += 1
    merge_rows(ws, 'A', 'N', idx)
    c = cell_center(
        ws,
        f'A{idx}',
        config.CHIEF_RANK + '_' * 75 + config.CHIEF_NAME)
    c.alignment = LEFT_ALIGNMENT
    c.font = Font(size=14)

    idx += 1
    merge_rows(ws, 'A', 'D', idx)
    date_fmt = '%d.%m.%Y'
    c = cell_center(ws, f'A{idx}', obj.document_date.strftime(date_fmt) + ' року')
    c.alignment = LEFT_ALIGNMENT
    c.font = Font(size=14)

    return idx


def format_rows(ws, data, fal_types):
    idx = 10
    fal_column = 6
    fal_names = [name for _, name in fal_types]
    for i, (k, d_data) in enumerate(data.items(), 1):
        rec_i = idx+i
        cell_center_border(ws, f'A{rec_i}', i)
        cell_center_border(ws, f'B{rec_i}', 'Донесення')
        cell_center_border(ws, f'C{rec_i}', k)
        cell_center_border(ws, f'D{rec_i}', d_data['document_date'])
        cell_center_border(ws, f'E{rec_i}', d_data['department'])
        for f_name, f_data in d_data['fals'].items():
            f_idx = fal_names.index(f_name)
            cell_center_border(ws,
                               f'{get_column_letter(f_idx*2+fal_column)}{rec_i}',
                               f_data['ltrs'])
            cell_center_border(ws,
                               f'{get_column_letter(1+f_idx*2+fal_column)}{rec_i}',
                               f_data['kgs'])
    return rec_i + 1


def format_table_header(ws, fal_types):
    idx = 6
    ranges = [
        f'A{idx}:A{idx+3}',
        f'B{idx}:B{idx+3}',
        f'C{idx}:C{idx+3}',
        f'D{idx}:D{idx+3}',
        f'E{idx}:E{idx+3}',
    ]
    for r in ranges:
        ws.merged_cells.ranges.add(r)
    fal_column = 4
    for _, ft in fal_types:
        fal_column += 2
        col_letter = get_column_letter(fal_column)

        ws.column_dimensions[col_letter].width = 24
        ws.column_dimensions[get_column_letter(fal_column+1)].width = 24

        ws.merged_cells.ranges.add(
            f'{
                col_letter}{idx+1}:{
                    get_column_letter(fal_column+1)}{idx+1}'
        )
        ws.merged_cells.ranges.add(
            f'{
                col_letter}{idx+2}:{
                    get_column_letter(fal_column+1)}{idx+2}'
        )

        cell_center_border(ws, f'{col_letter}{idx+1}', ft)
        cell_center_border(ws, f'{col_letter}{idx}', '')
        cell_center_border(ws, f'{get_column_letter(fal_column+1)}{idx}', '')
        cell_center_border(ws, f'{get_column_letter(fal_column+1)}{idx+1}', '')

        cell_center_border(ws, f'{col_letter}{idx+2}', 'ж/д')
        cell_center_border(ws, f'{get_column_letter(fal_column+1)}{idx+2}', '')
        cell_center_border(ws, f'{col_letter}{idx+3}', 'л')
        cell_center_border(ws, f'{get_column_letter(fal_column+1)}{idx+3}', 'кг')

    ws.merged_cells.ranges.add(f'F{idx}:{get_column_letter(fal_column+1)}{idx}')
    cell_center_border(ws, f'F{idx}', 'Найменування')

    cell_center_border(ws, f'A{idx}', '№ з/п')
    cell_vertical_border(ws, f'B{idx}', 'Найменування документа')
    cell_vertical_border(ws, f'C{idx}', '№ документа')
    cell_vertical_border(ws, f'D{idx}', 'Дата документа')
    cell_vertical_border(ws, f'E{idx}', 'Отримувач (здавальник)')
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16
    ws.row_dimensions[6].height = 20

    for i in range(1, fal_column+2):
        cell_center_border(ws, f'{get_column_letter(i)}{idx+4}', i)
