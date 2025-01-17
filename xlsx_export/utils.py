from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from dateutil.rrule import rrule, MONTHLY
from datetime import datetime

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical='center', wrap_text=True)
LEFT_ALIGNMENT = Alignment(vertical='center', wrap_text=True)
VERTICAL_ALIGNMENT = Alignment(
    textRotation=90,
    horizontal="center", vertical='center', wrap_text=True)

OTHER_DEP_CELL_FILL = PatternFill(
    start_color="dce6f1", end_color="dce6f1", fill_type="solid"
)
BASE_DEP_CELL_FILL = PatternFill(
    start_color="e1e7d5", end_color="e1e7d5", fill_type="solid"
)


def get_or_zero(ws, cn):
    return ws[cn].value or 0


def month_iter(start_month, start_year, end_month, end_year):
    start = datetime(start_year, start_month, 1)
    end = datetime(end_year, end_month, 1)

    return ((d.year, d.month) for d in rrule(MONTHLY, dtstart=start, until=end))


def cell_center(ws, cell_name, value):
    ws[cell_name] = value
    ws[cell_name].alignment = CENTER_ALIGNMENT
    return ws[cell_name]


def cell_center_border(ws, cell_name, value):
    c = cell_center(ws, cell_name, value)
    c.border = THIN_BORDER
    return c


def cell_vertical_border(ws, cell_name, value):
    c = cell_center_border(ws, cell_name, value)
    c.alignment = VERTICAL_ALIGNMENT
    return c


def header_cell_center_border(ws, cell_name, value):
    c = cell_center_border(ws, cell_name, value)
    c.font = Font(bold=True)
    c.fill = PatternFill(
        start_color="c4d79b", end_color="c4d79b", fill_type="solid"
    )


def merge_rows(ws, col_start, col_end, idx):
    ws.merged_cells.ranges.add(f'{col_start}{idx}:{col_end}{idx}')


def fill_empty_border(ws, rng, value=''):
    for cell in ws[rng]:
        for obj in cell:
            obj.border = THIN_BORDER
            if obj.value is None:
                obj.value = value
                obj.alignment = CENTER_ALIGNMENT
