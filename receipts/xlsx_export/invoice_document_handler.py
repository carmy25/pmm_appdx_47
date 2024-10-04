from openpyxl.utils import get_column_letter

from fals.models import FAL
from ..models.reporting import FALReportEntry
from .utils import CENTER_ALIGNMENT, THIN_BORDER, cell_center_border
from .base_document_handler import BaseFALDocumentHandler


class InvoiceDocumentHandler(BaseFALDocumentHandler):
    '''Process invoices'''

    def update_total_base_dep(self):
        pass

    def get_document_name(self):
        return self.fal.document_object._meta.verbose_name

    def get_dep(self):
        return ''

    def get_document_number(self):
        return self.fal.document_object.number or '-'

    def get_document_operation_date(self):
        return self.fal.document_object.operation_date

    def get_document_sender(self):
        return 'в підр.(нак)'

    def get_fal_income(self):
        return 0

    def get_fal_outcome(self):
        return 0

    def process(self):
        super().process()
        self.format_departments()
        return True

    def format_departments(self):
        doc = self.fal.document_object
        self.format_dep_fal(doc.sender, self.fal.amount, income=False)
        self.format_dep_fal(doc.destination, self.fal.amount)

    def format_dep_fal(self, dep, amount, income=True):
        dep_index = self.state['DEP_BY_INDEX'][dep.name]
        cell_center_border(
            self.ws,
            self.add_idx(
                get_column_letter(
                    dep_index if income else dep_index + 1)), amount)
        col_letter = get_column_letter(dep_index+2)
        total_cell = self.add_idx(col_letter)
        self.ws[total_cell].value = \
            f'=SUM({get_column_letter(dep_index)}$3:{get_column_letter(dep_index)}{self.state["idx"]})-SUM({
            get_column_letter(dep_index+1)}$3:{get_column_letter(dep_index+1)}{self.state["idx"]})'
        self.ws[total_cell].alignment = CENTER_ALIGNMENT
        self.ws[total_cell].border = THIN_BORDER
