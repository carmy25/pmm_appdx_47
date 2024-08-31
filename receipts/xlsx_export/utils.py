from openpyxl.styles import Border, \
    Side, PatternFill, Alignment

THIN_BORDER = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
CENTER_ALIGNMENT = Alignment(horizontal='center')

OTHER_DEP_CELL_FILL = PatternFill(
    start_color="dce6f1", end_color="dce6f1", fill_type="solid")
BASE_DEP_CELL_FILL = PatternFill(
    start_color="e1e7d5", end_color="e1e7d5", fill_type="solid")


def cell_center_border(ws, cell_name, value):
    ws[cell_name] = value
    ws[cell_name].alignment = CENTER_ALIGNMENT
    ws[cell_name].border = THIN_BORDER
    return ws[cell_name]
