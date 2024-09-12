from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from itertools import chain
from fals.models import FAL
from receipts.models.reporting import FALReportEntry
from ..models import ReceiptRequest, ReceiptRequestCoupon, Certificate

from .utils import CENTER_ALIGNMENT, cell_center_border, BASE_DEP_CELL_FILL, THIN_BORDER, OTHER_DEP_CELL_FILL


DEP_BY_INDEX = {}


class BaseFALDocumentHandler:
    def __init__(self, fal, ws, state):
        self.ws = ws
        self.fal = fal
        self.state = state

    def add_idx(self, col_name):
        return f"{col_name}{self.state['idx']}"

    def format_document_name(self, name):
        cell_center_border(self.ws, self.add_idx('A'), name)

    def format_document_number(self, number):
        cell_center_border(self.ws, self.add_idx('B'), number)

    def format_document_operation_date(self, date):
        cell_center_border(self.ws, self.add_idx('C'), date)

    def format_document_sender(self, sender):
        cell_center_border(self.ws, self.add_idx('D'), sender)

    def format_fal_income(self, amount):
        cell_center_border(self.ws, self.add_idx('E'), amount or '')

    def format_fal_income_dep(self, amount):
        cell_center_border(self.ws, self.add_idx('E'), amount or '')

    def format_fal_outcome(self, amount):
        cell_center_border(self.ws, self.add_idx('F'), amount or '')

    def format_fal_total(self):
        cell_center_border(self.ws, self.add_idx('G'), self.state['total'])

    def format_fal_by_dep(self):
        income = self.get_fal_income()
        outcome = self.get_fal_outcome()
        dep = self.get_dep()[1:]
        if dep == '4548':
            cell_center_border(self.ws, self.add_idx('H'), income or '')
            cell_center_border(self.ws, self.add_idx('I'), outcome or '')
            cell_center_border(self.ws, self.add_idx('K'), '')
            cell_center_border(self.ws, self.add_idx('L'), '')
        elif dep == '4635':
            cell_center_border(self.ws, self.add_idx('H'), '')
            cell_center_border(self.ws, self.add_idx('I'), '')
            cell_center_border(self.ws, self.add_idx('K'), income or '')
            cell_center_border(self.ws, self.add_idx('L'), outcome or '')

    def format_fal_total_base_dep(self):
        cell_center_border(self.ws, self.add_idx(
            'J'), self.state['total_by_dep']['4548'])
        cell_center_border(self.ws, self.add_idx(
            'M'), self.state['total_by_dep']['4635'])

    def process(self):
        self.format_document_name(self.get_document_name())
        self.format_document_number(self.get_document_number())
        self.format_document_operation_date(self.get_document_operation_date())
        self.format_document_sender(self.get_document_sender())
        self.format_fal_income(self.get_fal_income())
        self.format_fal_outcome(self.get_fal_outcome())

        self.update_total()
        self.format_fal_total()

        self.update_total_base_dep()
        self.format_fal_total_base_dep()
        self.format_fal_by_dep()

    def update_total(self):
        income = self.get_fal_income() or 0
        outcome = self.get_fal_outcome() or 0
        self.state['total'] = self.state['total'] - outcome + income


class ReportingSummaryReportDocumentHandler(BaseFALDocumentHandler):
    '''Handle ReportingSummaryReport
    '''

    def update_total_base_dep(self):
        pass

    def get_document_name(self):
        return self.fal.report.summary_report._meta.verbose_name

    def get_dep(self):
        return ''

    def get_document_number(self):
        return self.fal.report.summary_report.number or '-'

    def get_document_operation_date(self):
        return self.fal.report.summary_report.end_date

    def get_document_sender(self):
        return 'в підр.'

    def get_fal_income(self):
        return 0

    def get_fal_outcome(self):
        fals = FALReportEntry.objects.filter(
            fal_type=self.fal.fal_type,
            report__summary_report=self.fal.report.summary_report)
        return sum([fal.outcome for fal in fals])

    def process(self):
        if self.fal.report.summary_report in self.state.setdefault('reports_processed', []):
            return
        self.state['reports_processed'].append(
            self.fal.report.summary_report)
        super().process()
        self.format_departments()

    def format_departments(self):
        fals = FALReportEntry.objects.filter(
            fal_type=self.fal.fal_type,
            report__summary_report=self.fal.report.summary_report)
        for fal in fals:
            dep_index = DEP_BY_INDEX[fal.report.department.name]
            cell_center_border(self.ws, self.add_idx(
                get_column_letter(dep_index+1)), fal.outcome)
            col_letter = get_column_letter(dep_index+2)
            total_cell = self.add_idx(col_letter)
            self.ws[total_cell].value = \
                f'=SUM({get_column_letter(dep_index)}$3:{get_column_letter(dep_index)}{self.state["idx"]})-SUM({
                get_column_letter(dep_index+1)}$3:{get_column_letter(dep_index+1)}{self.state["idx"]})'
            self.ws[total_cell].alignment = CENTER_ALIGNMENT
            self.ws[total_cell].border = THIN_BORDER


class FALDocumentHandler(BaseFALDocumentHandler):

    def get_dep(self):
        if type(self.fal.document_object) in [ReceiptRequestCoupon, Certificate]:
            return self.fal.document_object.destination
        else:
            return self.fal.document_object.sender

    def update_total_base_dep(self):
        if type(self.fal.document_object) in [ReceiptRequestCoupon, Certificate]:
            self.state['total_by_dep'][self.fal.document_object.destination[1:]
                                       ] += self.fal.amount
        else:
            self.state['total_by_dep'][self.fal.document_object.sender[1:]
                                       ] -= self.fal.amount

    def get_document_name(self):
        return self.fal.document_object._meta.verbose_name

    def get_document_number(self):
        doc = self.fal.document_object
        if type(doc) in [ReceiptRequest, ReceiptRequestCoupon]:
            return f'{
                doc.number}/{doc.book_number}{doc.book_series.upper()}'
        return doc.number

    def get_fal_income(self):
        doc = self.fal.document_object
        if type(doc) in [Certificate, ReceiptRequestCoupon]:
            return self.fal.amount
        return 0

    def get_fal_outcome(self):
        doc = self.fal.document_object
        if type(doc) in [Certificate, ReceiptRequestCoupon]:
            return 0
        return self.fal.amount

    def get_document_sender(self):
        doc = self.fal.document_object
        if type(doc) in [Certificate, ReceiptRequestCoupon]:
            return doc.sender.upper()
        return doc.destination.upper()

    def get_document_operation_date(self):
        return self.fal.document_object.operation_date


def export_fal_type(fal_type, ws, departments):
    format_header(ws, fal_type, departments)
    format_rows(ws, fal_type)


def get_fal_date(obj):
    if type(obj) is FALReportEntry:
        return obj.report.summary_report.end_date
    return obj.document_object.operation_date


def get_sorted_fals(fal_type):
    fals = FAL.objects \
        .filter(fal_type=fal_type) \
        .exclude(object_id__isnull=True) \
        .prefetch_related('document_object')
    fal_report_entries = FALReportEntry.objects \
        .filter(fal_type=fal_type) \
        .exclude(report__isnull=True) \
        .exclude(report__summary_report__isnull=True)

    sorted_fals = sorted(
        chain(fals, fal_report_entries),
        key=get_fal_date)
    return sorted_fals


def format_rows(ws, fal_type):
    ws_state = {'total': 0,
                'total_by_dep': {'4548': 0, '4635': 0}}
    fals = get_sorted_fals(fal_type)
    for i, fal in enumerate(fals):
        ws_state['idx'] = i + 3
        if type(fal) is FALReportEntry:
            ReportingSummaryReportDocumentHandler(fal, ws, ws_state).process()
            continue
        FALDocumentHandler(fal, ws, ws_state).process()


def format_header(ws, fal_type, departments):
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['b'].width = 30
    ws.column_dimensions['c'].width = 30
    ws.column_dimensions['d'].width = 30
    ws.column_dimensions['e'].width = 20
    ws.column_dimensions['f'].width = 20
    ws.column_dimensions['g'].width = 20
    ws['A1'] = fal_type.name
    ws['a1'].border = THIN_BORDER
    ws[f'a1'].alignment = Alignment(horizontal='center')
    ws['a1'].font = Font(bold=True)
    ws['a1'].fill = PatternFill(
        start_color="c4d79b", end_color="c4d79b", fill_type="solid")
    ws.merged_cells.ranges.add('A1:D1')
    ws.merged_cells.ranges.add('E1:G1')
    total_cell = cell_center_border(ws, 'E1', 'Загалом')
    total_cell.font = Font(bold=True)
    total_cell.fill = PatternFill(
        start_color="fee8d9", end_color="fee8d9", fill_type="solid")

    cell_center_border(
        ws, 'a2', 'Найменування документу').font = Font(bold=True)
    cell_center_border(ws, 'b2', 'Номер документу').font = Font(bold=True)
    cell_center_border(ws, 'c2', 'Дата операції').font = Font(bold=True)
    cell_center_border(
        ws, 'd2', 'Постачальник (одержувач)').font = Font(bold=True)
    cell_center_border(ws, 'e2', 'Надійшло').font = Font(bold=True)
    cell_center_border(ws, 'f2', 'Вибуло').font = Font(bold=True)
    cell_center_border(ws, 'g2', 'Всього').font = Font(bold=True)

    base_dep_cell = cell_center_border(ws, 'H1', 'A4548')
    base_dep_cell.font = Font(bold=True)
    base_dep_cell.fill = BASE_DEP_CELL_FILL
    ws.merged_cells.ranges.add('H1:J1')
    ws[f'H1'].alignment = Alignment(horizontal='center')

    cell_center_border(ws, 'H2', 'Надійшло').font = Font(bold=True)
    cell_center_border(ws, 'I2', 'Вибуло').font = Font(bold=True)
    cell_center_border(ws, 'J2', 'Всього').font = Font(bold=True)

    other_dep_cell = cell_center_border(ws, 'K1', 'A4635')
    other_dep_cell.font = Font(bold=True)
    other_dep_cell.fill = OTHER_DEP_CELL_FILL
    ws.merged_cells.ranges.add('k1:m1')
    cell_center_border(ws, 'K2', 'Надійшло').font = Font(bold=True)
    cell_center_border(ws, 'L2', 'Вибуло').font = Font(bold=True)
    cell_center_border(ws, 'M2', 'Всього').font = Font(bold=True)

    format_departments(ws, departments)


def format_departments(ws, deps):
    col_idx = 14
    for dep in deps:
        DEP_BY_INDEX[dep.name] = col_idx
        start_cell_name = f'{get_column_letter(col_idx)}1'
        ws.merged_cells.ranges.add(f'{start_cell_name}:{
                                   get_column_letter(col_idx+2)}1')
        dep_name_cell = cell_center_border(ws, start_cell_name, dep.name)
        dep_name_cell.font = Font(bold=True)
        dep_name_cell.fill = PatternFill(
            start_color="fee8d9", end_color="fee8d9", fill_type="solid")

        cell_center_border(ws, f'{get_column_letter(
            col_idx)}2', 'Надійшло').font = Font(bold=True)
        cell_center_border(ws, f'{get_column_letter(
            col_idx+1)}2', 'Вибуло').font = Font(bold=True)
        cell_center_border(ws, f'{get_column_letter(
            col_idx+2)}2', 'Всього').font = Font(bold=True)
        col_idx += 3
