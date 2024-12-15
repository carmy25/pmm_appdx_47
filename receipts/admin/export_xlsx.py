from time import ctime
from django.urls import reverse
import pandas as pd

from django.db.models import F
from django.contrib import admin
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.utils.safestring import mark_safe
from django.contrib import messages
from django.conf import settings


from departments.models import Department
from fals.models import FALType
from receipts.exceptions import DuplicateReportsError
from receipts.forms import StocktakingSettingsForm
from receipts.models.receipt import InvoiceForRRC
from receipts.models.reporting import Reporting
from xlsx_export.invoices_for_rrc import InvoiceForRRCMut, format_price_summary
from xlsx_export.reportings_report import export_reportings_report
from xlsx_export import export_fal_type
from xlsx_export.stocktracking_report import export_stocktaking_report
from xlsx_export.utils import month_iter

import openpyxl

from xlsx_export.xlsx_export import export_reportings_fes_registry, export_reportings_price_report


def xlsx_response(filename):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={filename}.xlsx"
    return response


@admin.site.register_view("export-xlsx", "47 Додаток", urlname="export_xlsx")
def export_xlsx(request):
    import time

    start = time.time()
    wb = openpyxl.Workbook()
    category = request.GET.get("category")
    fal_types = (
        FALType.objects.filter(category=category) if category else FALType.objects.all()
    )
    departments = (
        Department.objects.all()
        .order_by("-name")
        .order_by(F("warehouse__order").asc(nulls_last=True))
    )
    for fal_type in fal_types:
        ws = wb.create_sheet(fal_type.name.replace("/", " ").replace(':', '')[:30])
        export_fal_type(fal_type, ws, departments)
        ws.freeze_panes = ws["B3"]

    end = time.time()
    print(end - start)
    response = xlsx_response("47appendix")
    wb.save(response)
    print(time.time() - end)
    return response


@admin.site.register_view(
    "reportings-report", "Звіт по донесеннях", urlname="reportings_report"
)
def reportings_report(request):
    response = xlsx_response("reportings_report")
    reportings = Reporting.objects.all().order_by("end_date")
    wb = openpyxl.Workbook()
    ws = wb['Sheet']
    ws.title = 'Звіт по донесеннях'
    if reportings.count() > 0:
        try:
            export_reportings_report(ws, reportings)
        except DuplicateReportsError as e:
            return render(
                request,
                "admin/reportings_report_error.html",
                context={"error_msg": e.args[0]},
            )
    wb.save(response)
    return response


@admin.site.register_view(
    "fes-registry", "Реєстр ФЕС", urlname="fes_registry"
)
def reportings_report(request):
    response = xlsx_response("fes_registry")
    reportings = Reporting.objects.all().order_by("end_date")
    wb = openpyxl.Workbook()
    if (fr := reportings.first()) is None:
        return response
    lr = reportings.last()

    for d in month_iter(
            fr.end_date.month, fr.end_date.year,
            lr.end_date.month, lr.end_date.year):
        year, month = d
        ws = wb.create_sheet(f'{month}.{year}')
        export_reportings_fes_registry(
            ws,
            reportings.filter(end_date__year=year, end_date__month=month),
            d)
        ws.freeze_panes = ws['A3']
    wb.save(response)
    return response


@admin.site.register_view(
    "reporting-price-report", "Звіт списання по донесеннях(в ціні)", urlname="reporting_price_report"
)
def reporting_price_report(request):
    response = xlsx_response("reporting_price_report")
    reportings = Reporting.objects.all().order_by("end_date")
    wb = openpyxl.Workbook()
    if (fr := reportings.first()) is None:
        return response
    lr = reportings.last()
    # Get all rrc invoices and sort by date
    invoices = sorted(InvoiceForRRC.objects.all(),
                      key=lambda x:
                          (x.rrc and x.rrc.operation_date) or
                          (x.certificate and x.certificate.operation_date))
    # convert to InvoiceForRRCMut
    invoices_for_rrc_mut = [InvoiceForRRCMut(inv) for inv in invoices]
    # put date to pandas DataFrame
    df = pd.DataFrame({
        'date': [inv.operation_date for inv in invoices_for_rrc_mut],
        'invoices': invoices_for_rrc_mut,
    })
    df['date'] = pd.to_datetime(df['date'])
    invoices_df = df.set_index('date')
    if len(invoices) == 0:
        return response
    inv = invoices[0]
    start_date = inv.operation_date
    months = list(month_iter(
        fr.end_date.month, fr.end_date.year,
        lr.end_date.month, lr.end_date.year))

    for d in months:
        year, month = d
        if year == 2023 and month < 6:
            continue
        ws = wb.create_sheet(f'{month}.{year}')
        export_reportings_price_report(
            ws,
            reportings.filter(end_date__year=year, end_date__month=month),
            d,
            invoices_df,
            start_date)
        ws.freeze_panes = ws['A3']

    ws = wb.create_sheet('Підсумок')
    format_price_summary(ws, months)

    wb.save(response)
    return response


@admin.site.register_view(
    'stocktaking-report', 'Інвентиризація',
    urlname='stocktaking_report'
)
def stocktaking_report(request):
    if request.method == "POST":
        # create a form instance and populate it with data from the request:
        form = StocktakingSettingsForm(request.POST)
        # check whether it's valid:
        if form.is_valid():
            wb = openpyxl.Workbook()
            departments = (
                Department.objects.all()
                .order_by("-name")
                .order_by(F("warehouse__order").asc(nulls_last=True))
            )
            export_stocktaking_report(wb, departments, form.cleaned_data)
            file_path = settings.MEDIA_ROOT / f'stocktaking-{ctime()}.xlsx'
            wb.save(filename=file_path)
            file_url = settings.MEDIA_URL + file_path.name

            messages.add_message(request,
                                 messages.INFO,
                                 mark_safe(
                                     f'<a href="{file_url}">Звіт по інвентаризації</a> створено. <a href="{file_url}">ЗАВАНТАЖИТИ</a>'))

            return HttpResponseRedirect(reverse('admin:index'))

            # if a GET (or any other method) we'll create a blank form
    else:
        form = StocktakingSettingsForm()

    return render(
        request,
        "admin/stocktaking_settings.html",
        {"form": form},
    )
