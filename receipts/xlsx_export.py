from openpyxl.styles import Border, \
    Side, PatternFill, Font, GradientFill, Alignment
import logging
from fals.models import FAL
from .models import ReceiptRequest, ReceiptRequestCoupon, Certificate

THIN_BORDER = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
CENTER_ALIGNMENT = Alignment(horizontal='center')

OTHER_DEP_CELL_FILL = PatternFill(
    start_color="dce6f1", end_color="dce6f1", fill_type="solid")
BASE_DEP_CELL_FILL = PatternFill(
    start_color="e1e7d5", end_color="e1e7d5", fill_type="solid")


def cell_center_border(ws, cell_name, value):
    ws[cell_name] = value
    ws[cell_name].alignment = CENTER_ALIGNMENT
    ws[cell_name].border = THIN_BORDER
    return ws[cell_name]


def export_fal_type(fal_type, ws):
    format_header(ws, fal_type)
    format_rows(ws, fal_type)


def format_rows(ws, fal_type):
    fals = FAL.objects.filter(fal_type=fal_type).order_by(
        'document__operation_date')
    total = 0
    total_by_dep = {'4548': 0, '4635': 0}
    for i, fal in enumerate(sorted(filter(lambda x: x.document_object, fals), key=lambda x: x.document_object.operation_date)):
        document = fal.document_object
        idx = i + 3
        cell_center_border(ws, f'A{idx}', document._meta.verbose_name)
        if type(document) in [ReceiptRequestCoupon, ReceiptRequest]:
            cell_center_border(ws, f'B{idx}', f'{
                document.number}/{document.book_number}{document.book_series}')
        else:
            cell_center_border(ws, f'B{idx}', document.number)
        cell_center_border(ws, f'C{idx}', document.operation_date)
        if type(document) in [ReceiptRequestCoupon, Certificate]:
            cell_center_border(ws, f'D{idx}', document.sender)
            amount_cell_range = f'E{idx}'
            amount_dep_cell_range = f'H{idx}' \
                if document.destination[1:] == '4548' else f'K{idx}'
        else:
            cell_center_border(ws, f'D{idx}', document.destination)
            amount_cell_range = f'F{idx}'
            if type(document) in [ReceiptRequestCoupon, Certificate]:
                amount_dep_cell_range = f'I{idx}' \
                    if document.destination[1:] == '4548' else f'L{idx}'
            else:
                amount_dep_cell_range = f'I{idx}' \
                    if document.sender[1:] == '4548' else f'L{idx}'

        cell_center_border(ws, amount_cell_range, fal.amount)
        cell_center_border(ws, amount_dep_cell_range, fal.amount)

        if type(document) in [ReceiptRequestCoupon, Certificate]:
            total += fal.amount
            total_by_dep[document.destination[1:]] += fal.amount
        else:
            total -= fal.amount
            total_by_dep[document.sender[1:]] -= fal.amount

        cell_center_border(ws, f'G{idx}', total)

        cell_center_border(ws, f'J{idx}', total_by_dep['4548'])
        cell_center_border(ws, f'M{idx}', total_by_dep['4635'])


def format_header(ws, fal_type):
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['b'].width = 30
    ws.column_dimensions['c'].width = 30
    ws.column_dimensions['d'].width = 30
    ws.column_dimensions['e'].width = 20
    ws.column_dimensions['f'].width = 20
    ws.column_dimensions['g'].width = 20
    ws['A1'] = fal_type.name
    ws['a1'].border = THIN_BORDER
    ws[f'a1'].alignment = Alignment(horizontal='center')
    ws['a1'].font = Font(bold=True)
    ws['a1'].fill = PatternFill(
        start_color="c4d79b", end_color="c4d79b", fill_type="solid")
    ws.merged_cells.ranges.add('A1:D1')
    ws.merged_cells.ranges.add('E1:G1')
    total_cell = cell_center_border(ws, 'E1', 'Загалом')
    total_cell.font = Font(bold=True)
    total_cell.fill = PatternFill(
        start_color="fee8d9", end_color="fee8d9", fill_type="solid")

    cell_center_border(
        ws, 'a2', 'Найменування документу').font = Font(bold=True)
    cell_center_border(ws, 'b2', 'Номер документу').font = Font(bold=True)
    cell_center_border(ws, 'c2', 'Дата операції').font = Font(bold=True)
    cell_center_border(
        ws, 'd2', 'Постачальник (одержувач)').font = Font(bold=True)
    cell_center_border(ws, 'e2', 'Надійшло').font = Font(bold=True)
    cell_center_border(ws, 'f2', 'Вибуло').font = Font(bold=True)
    cell_center_border(ws, 'g2', 'Всього').font = Font(bold=True)

    base_dep_cell = cell_center_border(ws, 'H1', 'A4548')
    base_dep_cell.font = Font(bold=True)
    base_dep_cell.fill = BASE_DEP_CELL_FILL
    ws.merged_cells.ranges.add('H1:J1')
    ws[f'H1'].alignment = Alignment(horizontal='center')

    cell_center_border(ws, 'H2', 'Надійшло').font = Font(bold=True)
    cell_center_border(ws, 'I2', 'Вибуло').font = Font(bold=True)
    cell_center_border(ws, 'J2', 'Всього').font = Font(bold=True)

    other_dep_cell = cell_center_border(ws, 'K1', 'A4635')
    other_dep_cell.font = Font(bold=True)
    other_dep_cell.fill = OTHER_DEP_CELL_FILL
    ws.merged_cells.ranges.add('k1:m1')
    cell_center_border(ws, 'K2', 'Надійшло').font = Font(bold=True)
    cell_center_border(ws, 'L2', 'Вибуло').font = Font(bold=True)
    cell_center_border(ws, 'M2', 'Всього').font = Font(bold=True)
