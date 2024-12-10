from django.core.management.base import BaseCommand
import openpyxl
from pprint import pprint


class Command(BaseCommand):
    help = "Check diff for waybills"

    def add_arguments(self, parser):
        parser.add_argument("file")

    def handle(self, *args, **options):
        wb = openpyxl.load_workbook(options['file'])
        ws = wb.active
        empty = []
        ok = []
        bad = []
        reason = {}
        for i in range(5, 52):
            orig = self.to_set(ws[f'E{i}'].value)
            facade = self.to_set(ws[f'F{i}'].value)
            waybills = self.to_set(ws[f'G{i}'].value)
            dep_name = ws[f'C{i}'].value

            if orig == facade == waybills:
                if orig == set():
                    empty.append(dep_name)
                    continue
                ok.append(dep_name)
                continue
            bad.append(dep_name)
            reason[dep_name] = (orig | facade | waybills) - (orig & facade & waybills)
        self.print_list(ok, '== Шляхові співпадаюь ==',)
        self.print_list(empty, '== Не вказані шляхові ==',)
        self.print_list(bad, '== Шляхові не співпадають ==', reason)

    def print_list(self, lst, title, reason=None):
        print(title)
        for dep in lst:
            print(f'\t{dep}\t')
            if (r := (reason is not None) and reason.get(dep)):
                pprint(r)

    def to_set(self, v):
        if v is None:
            return set()
        return {x.strip() for x in v.split(',')}
