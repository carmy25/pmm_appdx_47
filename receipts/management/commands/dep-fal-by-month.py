import shutil
from django.core.management.base import BaseCommand
from departments.models import Department
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from datetime import datetime, timedelta, date

from receipts.models.handout_list import HandoutList


class Command(BaseCommand):
    help = "Create excel file for department with received fals by month"

    def add_arguments(self, parser):
        parser.add_argument("start")
        parser.add_argument("end")
        parser.add_argument('dep_name')
        parser.add_argument("dest")

    def handle(self, *args, **options):
        wb = Workbook()
        ws = wb.active
        dep_name = options['dep_name']
        ws.title = dep_name
        start = options['start']
        end = options['end']

        # Define first column values
        first_column_values = ['a-80', 'a-92', 'a-95', 'ДП']
        for row, value in enumerate(first_column_values, start=2):  # Start from row 2
            ws.cell(row=row, column=1, value=value)

        department = Department.objects.filter(name=dep_name).first()

        for r in HandoutList.objects.filter(
                operation_date__gt=start,
                operation_date__lt=end,
                destination=department):
            for fal in r.fals.all():
                fal_type = fal.fal_type.name
                amount = fal.amount

        # Define border style
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Apply borders to the range A1:G4
        for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=7):
            for cell in row:
                cell.border = thin_border

        dest = options['dest']
        wb.save(dest)
        print(f"Excel file '{dest}' has been created successfully.")
